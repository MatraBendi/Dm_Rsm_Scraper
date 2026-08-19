"""Excel export (openpyxl) – Termékek, Összesítő és Info munkalap."""
from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from statistics import mean
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import Product, RunStats

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS: List[Tuple[str, str, int]] = [
    ("scraped_date", "Dátum", 12),
    ("store", "Bolt", 11),
    ("brand_group", "Kategória", 24),
    ("brand", "Márka", 20),
    ("name", "Terméknév", 60),
    ("size_label", "Kiszerelés", 14),
    ("size_value", "Mennyiség", 12),
    ("size_unit", "Egység", 9),
    ("pack_count", "Csomag (db)", 12),
    ("price", "Ár (Ft)", 12),
    ("old_price", "Akció előtti ár", 15),
    ("unit_price", "Egységár", 16),
    ("ean", "EAN", 16),
    ("sku", "Cikkszám", 14),
    ("availability", "Elérhetőség", 18),
    ("url", "Termék URL", 55),
    ("scraped_at", "Időbélyeg", 20),
    ("extraction", "Forrás", 22),
]


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def write_products_sheet(wb: Workbook, products: List[Product]) -> None:
    ws = wb.active
    ws.title = "Termékek"

    for idx, (_, title, width) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=title)
        ws.column_dimensions[get_column_letter(idx)].width = width
    _style_header(ws, 1, len(COLUMNS))

    for r, prod in enumerate(products, start=2):
        data = prod.to_dict()
        for c, (key, _, _w) in enumerate(COLUMNS, start=1):
            value = data.get(key)
            cell = ws.cell(row=r, column=c)
            if key == "ean" and value:
                cell.value = str(value)
                cell.number_format = "@"
            elif key in ("price", "old_price") and isinstance(value, (int, float)):
                cell.value = float(value)
                cell.number_format = '# ##0 "Ft"'
            elif key == "size_value" and isinstance(value, (int, float)):
                cell.value = float(value)
                cell.number_format = "0.##"
            elif key == "url" and value:
                cell.value = str(value)
                cell.hyperlink = str(value)
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.value = value if value is not None else ""
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(key == "name"))

    last_row = max(2, len(products) + 1)
    ref = "A1:{}{}".format(get_column_letter(len(COLUMNS)), last_row)
    try:
        table = Table(displayName="Termekek", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
        ws.add_table(table)
    except Exception:  # noqa: BLE001 – üres adathalmaznál kihagyjuk
        ws.auto_filter.ref = ref
    ws.freeze_panes = "A2"


def write_summary_sheet(wb: Workbook, products: List[Product]) -> None:
    ws = wb.create_sheet("Összesítő")
    stores = sorted({p.store for p in products}) or ["dm", "Rossmann"]

    ws.cell(row=1, column=1, value="Márka × kiszerelés összesítő").font = TITLE_FONT
    header_row = 3
    headers = ["Kategória", "Márka", "Kiszerelés"]
    for store in stores:
        headers += ["{} – db".format(store), "{} – átlagár".format(store),
                    "{} – min".format(store), "{} – max".format(store)]
    headers += ["Árkülönbség (max-min bolt között)"]

    for idx, title in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=title)
        ws.column_dimensions[get_column_letter(idx)].width = 18 if idx > 3 else 26
    _style_header(ws, header_row, len(headers))

    buckets: Dict[Tuple[str, str, str], Dict[str, List[float]]] = defaultdict(lambda: defaultdict(list))
    for p in products:
        if p.price is None:
            continue
        key = (p.brand_group, p.brand, p.size_label or "n/a")
        buckets[key][p.store].append(float(p.price))

    r = header_row + 1
    for key in sorted(buckets.keys(), key=lambda k: (k[0], k[1], _size_sort(k[2]))):
        group, brand, size = key
        ws.cell(row=r, column=1, value=group)
        ws.cell(row=r, column=2, value=brand)
        ws.cell(row=r, column=3, value=size)
        col = 4
        store_avgs: List[float] = []
        for store in stores:
            vals = buckets[key].get(store, [])
            if vals:
                ws.cell(row=r, column=col, value=len(vals))
                for offset, val in enumerate([mean(vals), min(vals), max(vals)], start=1):
                    cell = ws.cell(row=r, column=col + offset, value=round(val, 1))
                    cell.number_format = '# ##0 "Ft"'
                store_avgs.append(mean(vals))
            col += 4
        if len(store_avgs) > 1:
            cell = ws.cell(row=r, column=col, value=round(max(store_avgs) - min(store_avgs), 1))
            cell.number_format = '# ##0 "Ft"'
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    ws.freeze_panes = "A{}".format(header_row + 1)
    if r > header_row + 1:
        ws.auto_filter.ref = "A{}:{}{}".format(header_row, get_column_letter(len(headers)), r - 1)


def _size_sort(label: str) -> Tuple[float, str]:
    import re
    m = re.search(r"(\d+(?:[.,]\d+)?)", label or "")
    return (float(m.group(1).replace(",", ".")) if m else 1e9, label or "")


def write_info_sheet(wb: Workbook, stats: RunStats, products: List[Product]) -> None:
    ws = wb.create_sheet("Info")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    rows = [
        ("Készült", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Futás kezdete", stats.started_at),
        ("Futás vége", stats.finished_at or ""),
        ("Boltok", ", ".join(stats.stores)),
        ("Márkák", ", ".join(stats.brands)),
        ("Termékek száma", len(products)),
        ("Ebből árral", sum(1 for p in products if p.price is not None)),
        ("Ebből EAN-nal", sum(1 for p in products if p.ean)),
        ("", ""),
        ("Megjegyzés", "Az árak a webshopok online listaárai a fenti időpontban. "
                       "Az üzletekben eltérő ár lehet érvényben."),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    start = len(rows) + 2
    ws.cell(row=start, column=1, value="Figyelmeztetések / hibák").font = TITLE_FONT
    r = start + 1
    for msg in (stats.warnings + stats.errors) or ["nincs"]:
        ws.cell(row=r, column=1, value=msg)
        r += 1


def export_excel(products: List[Product], stats: RunStats, path: str) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    products = sorted(products, key=lambda p: (p.brand_group, p.brand, _size_sort(p.size_label), p.store, p.name))
    wb = Workbook()
    write_products_sheet(wb, products)
    write_summary_sheet(wb, products)
    write_info_sheet(wb, stats, products)
    wb.save(path)
    return path


def default_filename(stores: List[str], folder: Optional[str] = None) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    name = "PG_arak_{}_{}.xlsx".format("-".join(s.lower() for s in stores) or "all", stamp)
    return os.path.join(folder or "kimenet", name)
